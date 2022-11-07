import openpyxl

name_file = ['', 'perso1', 'person2', 'person3', 'person4', 'person5']
file_1 = ['id', '123456', '789654', '321465', '135802', '133456']
file_2 = ['firstname', 'name', 'name1', 'name2', 'name3', 'name4']
file_3 = ['age', 22, 25, 30, 32, 38]


wb = openpyxl.Workbook()
wb.create_sheet(title='Первый лист', index=0)
wb.remove(wb['Sheet'])
sheet = wb['Первый лист']

for row_index, row in enumerate((name_file, file_1, file_2, file_3)):
    for col_index, value in enumerate(row):
        cell = sheet.cell(row=row_index + 1, column=col_index + 1)
        cell.value = value
wb.save('task_3.xlsx')

